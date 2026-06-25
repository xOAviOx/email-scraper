"""
Format a scraped leads CSV into a professional, outreach-ready Excel workbook.

Reads the CSV produced by maps_email_scraper.py (category, location, name,
website, phone, address, rating, emails) and writes a styled .xlsx with:
  * a cleaned, validated "Leads" sheet (deduped, emails split out, rows with
    an email highlighted), and
  * a "Summary" sheet with per-category / per-location / per-tier analytics.

Usage:
    python format_leads.py
    python format_leads.py --input leads_india.csv --output leads_india_formatted.xlsx
"""

import argparse
import csv
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Kept in sync with maps_email_scraper.py so we can tag each city's tier.
TIER1 = {
    "Mumbai", "Delhi", "Bangalore", "Hyderabad",
    "Ahmedabad", "Chennai", "Kolkata", "Pune",
}

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
JUNK_EXTENSIONS = (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".css",
                   ".js", ".ico", ".woff", ".woff2", ".ttf", ".mp4", ".pdf",
                   ".webm")
JUNK_DOMAINS = ("example.com", "example.org", "email.com", "domain.com",
                "yourdomain.com", "sentry.io", "wixpress.com",
                "sentry-next.wixpress.com", "mysite.com", "company.com",
                "godaddy.com", "placeholder.com")


def clean_email(raw: str) -> str:
    """Strip encoding corruption / non-printable bytes and validate one email."""
    if not raw:
        return ""
    # Remove anything that isn't a normal email character (kills the stray
    # control bytes that show up when a site's encoding was mangled).
    e = "".join(ch for ch in raw if 32 < ord(ch) < 127)
    e = e.lower().strip().strip(".")
    m = EMAIL_RE.search(e)
    if not m:
        return ""
    e = m.group()
    local, _, domain = e.rpartition("@")
    if not local or "." not in domain:
        return ""
    if e.endswith(JUNK_EXTENSIONS):
        return ""
    if any(domain == d or domain.endswith("." + d) for d in JUNK_DOMAINS):
        return ""
    if len(e) > 60 or len(local) > 40:
        return ""
    return e


def clean_emails(cell: str) -> list[str]:
    out = []
    for part in re.split(r"[;,\s]+", cell or ""):
        e = clean_email(part)
        if e and e not in out:
            out.append(e)
    return out


def tier_of(city: str) -> str:
    return "Tier 1" if city in TIER1 else "Tier 2"


def load_rows(path: str) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    cleaned, seen = [], set()
    for r in rows:
        name = (r.get("name") or "").strip()
        if not name:
            continue
        website = (r.get("website") or "").strip()
        emails = clean_emails(r.get("emails", ""))
        # Dedupe on website, else name+phone.
        phone = (r.get("phone") or "").strip()
        key = website.lower().rstrip("/") if website else f"{name.lower()}|{phone}"
        if key in seen:
            continue
        seen.add(key)
        city = (r.get("location") or "").strip()
        cleaned.append({
            "category": (r.get("category") or "").strip(),
            "location": city,
            "tier": tier_of(city),
            "name": name,
            "website": website,
            "phone": phone,
            "email": "; ".join(emails),
            "n_emails": len(emails),
            "address": (r.get("address") or "").strip(),
            "rating": (r.get("rating") or "").strip(),
        })
    return cleaned


# --- styling helpers -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EMAIL_FILL = PatternFill("solid", fgColor="E2EFDA")      # green = has email
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top")


def build_workbook(rows: list[dict], out_path: str) -> None:
    wb = Workbook()
    _leads_sheet(wb.active, rows)
    _summary_sheet(wb.create_sheet("Summary"), rows)
    wb.save(out_path)


def _leads_sheet(ws, rows):
    ws.title = "Leads"
    headers = ["Category", "Location", "Tier", "Business Name", "Website",
               "Phone", "Email(s)", "Address", "Rating"]
    keys = ["category", "location", "tier", "name", "website", "phone",
            "email", "address", "rating"]
    widths = [20, 18, 8, 34, 38, 20, 34, 46, 8]

    ws.append(headers)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Sort: leads with emails first, then by category, location, name.
    rows = sorted(rows, key=lambda r: (r["n_emails"] == 0, r["category"],
                                       r["location"], r["name"].lower()))
    for r in rows:
        ws.append([r[k] for k in keys])
        row_cells = ws[ws.max_row]
        for cell in row_cells:
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        row_cells[2].alignment = CENTER   # Tier
        row_cells[8].alignment = CENTER   # Rating
        if r["n_emails"]:
            for cell in row_cells:
                cell.fill = EMAIL_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _summary_sheet(ws, rows):
    total = len(rows)
    with_email = sum(1 for r in rows if r["n_emails"])
    with_site = sum(1 for r in rows if r["website"])
    with_phone = sum(1 for r in rows if r["phone"])

    def pct(n):
        return f"{(n / total * 100):.1f}%" if total else "0%"

    ws["A1"] = "India Tier 1 + Tier 2 — Leads Summary"
    ws["A1"].font = TITLE_FONT
    ws.append([])

    def section(title, pairs, headers=("", "Count", "% of total")):
        ws.append([title])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E78")
        ws.append(list(headers))
        for cell in ws[ws.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for label, n in pairs:
            ws.append([label, n, pct(n)])
        ws.append([])

    section("Overview", [
        ("Total unique leads", total),
        ("Leads with an email", with_email),
        ("Leads with a website", with_site),
        ("Leads with a phone", with_phone),
    ])

    def group(field):
        agg = {}
        for r in rows:
            k = r[field] or "(unknown)"
            a = agg.setdefault(k, [0, 0])
            a[0] += 1
            a[1] += 1 if r["n_emails"] else 0
        return agg

    # By tier
    tier = group("tier")
    ws.append(["By Tier"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E78")
    ws.append(["Tier", "Leads", "With email"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for k in sorted(tier):
        ws.append([k, tier[k][0], tier[k][1]])
    ws.append([])

    # By category
    cat = group("category")
    ws.append(["By Category"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E78")
    ws.append(["Category", "Leads", "With email"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for k in sorted(cat, key=lambda x: -cat[x][0]):
        ws.append([k, cat[k][0], cat[k][1]])
    ws.append([])

    # By location
    loc = group("location")
    ws.append(["By Location"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12, color="1F4E78")
    ws.append(["Location", "Leads", "With email"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for k in sorted(loc, key=lambda x: -loc[x][0]):
        ws.append([k, loc[k][0], loc[k][1]])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="leads_india.csv")
    ap.add_argument("--output", default="leads_india_formatted.xlsx")
    args = ap.parse_args()

    rows = load_rows(args.input)
    build_workbook(rows, args.output)
    with_email = sum(1 for r in rows if r["n_emails"])
    print(f"Wrote {len(rows)} leads to {args.output} ({with_email} with emails)")


if __name__ == "__main__":
    main()
